"""
Excel Export – xuất tất cả báo cáo ra file .xlsx
Định dạng chuyên nghiệp, phù hợp nộp kiểm toán GRS/GOTS

Sheets được tạo:
  1. Annual_MB        – Annual Mass Balance (= "Xuat nhap Ton")
  2. Monthly_MB       – Monthly breakdown table
  3. Summary          – Summary Maintain Stock (= "Summary")
  4. Stock_Card       – Thẻ kho theo vật liệu
  5. Traceability     – Truy xuất nguồn gốc theo In TC
  6. Loss_Analysis    – Phân tích hao hụt
  7. TC_In_List       – Danh sách In TC
  8. TC_Out_List      – Danh sách Out TC
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
    numbers as xl_numbers,
)
from openpyxl.utils import get_column_letter

from app.core.mass_balance import MassBalanceResult
from app.reports.loss_analysis import LossAnalysisResult
from app.reports.stock_card import StockCardResult, StockSummaryRow
from app.reports.traceability import TraceabilityResult


# ─────────────────────────────────────────────────────────────────────────────
# Style constants / Hằng số định dạng
# ─────────────────────────────────────────────────────────────────────────────

# Màu sắc GRS-compliant / GRS-compliant color scheme
C_HEADER_BG    = "1F4E79"   # Xanh đậm / Dark blue
C_HEADER_FG    = "FFFFFF"   # Chữ trắng / White text
C_SUBHEADER_BG = "2E75B6"   # Xanh vừa / Medium blue
C_SECTION_BG   = "BDD7EE"   # Xanh nhạt / Light blue (section titles)
C_ALT_ROW      = "F2F7FB"   # Xanh rất nhạt cho dòng chẵn / Alt row
C_TOTAL_BG     = "FFF2CC"   # Vàng nhạt cho dòng tổng / Totals row
C_PASS         = "E2EFDA"   # Xanh lá nhạt / PASS green
C_WARNING      = "FFEB9C"   # Vàng nhạt / WARNING yellow
C_FAIL         = "FFC7CE"   # Đỏ nhạt / FAIL red
C_DIFF_BG      = "FFFACD"   # Vàng kem / Cream for difference row

# Fonts
F_TITLE   = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
F_HEADER  = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
F_BOLD    = Font(name="Calibri", bold=True, size=10)
F_NORMAL  = Font(name="Calibri", size=10)
F_SMALL   = Font(name="Calibri", size=9, color="595959")

# Number formats
FMT_KG    = '#,##0.000 "kg"'
FMT_PCT   = '0.00"%"'
FMT_PCT4  = '0.0000"%"'
FMT_DATE  = 'DD-MMM-YYYY'
FMT_INT   = '#,##0'


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border(style="thin") -> Border:
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


# ─────────────────────────────────────────────────────────────────────────────
# Helper: apply header row
# ─────────────────────────────────────────────────────────────────────────────

def _write_header_row(ws, row: int, headers: List[str], widths: Optional[List[int]] = None) -> None:
    """Ghi dòng tiêu đề bảng với định dạng đậm / Write formatted table header row"""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = F_HEADER
        cell.fill      = _fill(C_SUBHEADER_BG)
        cell.alignment = _center()
        cell.border    = _border()
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


def _write_title_block(ws, row: int, title: str, subtitle: str = "", company: str = "") -> int:
    """Ghi block tiêu đề trang / Write page title block"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    t = ws.cell(row=row, column=1, value=title)
    t.font      = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    t.alignment = _center()

    if subtitle:
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=10)
        s = ws.cell(row=row+1, column=1, value=subtitle)
        s.font      = Font(name="Calibri", size=11, color="595959")
        s.alignment = _center()

    if company:
        ws.merge_cells(start_row=row+2, start_column=1, end_row=row+2, end_column=10)
        c = ws.cell(row=row+2, column=1, value=company)
        c.font      = Font(name="Calibri", size=10, color="595959")
        c.alignment = _center()

    return row + 3 + 1   # Dòng bắt đầu nội dung / Content start row


def _status_fill(status: str) -> PatternFill:
    return _fill({
        "PASS"    : C_PASS,
        "WARNING" : C_WARNING,
        "FAIL"    : C_FAIL,
    }.get(status, "FFFFFF"))


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Annual Mass Balance  (= "Xuat nhap Ton")
# ─────────────────────────────────────────────────────────────────────────────

def write_annual_mass_balance(
    ws,
    result    : MassBalanceResult,
    company   : str = "",
) -> None:
    """
    Ghi báo cáo Mass Balance năm vào worksheet
    Write annual Mass Balance report to worksheet

    Format khớp với sheet "Xuat nhap Ton" trong file mẫu
    Format matches "Xuat nhap Ton" sheet in the sample file
    """
    ws.title = "Annual_Mass_Balance"
    ws.sheet_view.showGridLines = False

    # ── Title block ───────────────────────────────────────────────────────────
    row = 1
    ws.merge_cells(f"A{row}:J{row}")
    t = ws.cell(row=row, column=1, value="ANNUAL MASS BALANCE REPORT / BÁO CÁO MASS BALANCE NĂM")
    t.font      = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    t.alignment = _center()

    row = 2
    ws.merge_cells(f"A{row}:J{row}")
    ws.cell(row=row, column=1, value=(
        f"Certification Standard: {result.certification_standard}  |  "
        f"Period: {result.period_start} → {result.period_end}  |  "
        f"Material: {result.material_code or 'All'}"
    )).alignment = _center()

    if company:
        row = 3
        ws.merge_cells(f"A{row}:J{row}")
        ws.cell(row=row, column=1, value=f"Factory: {company}").alignment = _center()

    row = 5

    # ── Helper để ghi một dòng label/value ────────────────────────────────────
    def write_row(label, value, note="", section_break=False, total_row=False, highlight=None):
        nonlocal row
        if section_break:
            row += 1

        # Label (cột A-C)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        lc = ws.cell(row=row, column=1, value=label)
        lc.font   = F_BOLD if total_row else F_NORMAL
        lc.border = _border("thin")

        # Value (cột D-F)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        vc = ws.cell(row=row, column=4, value=value)
        vc.number_format = FMT_KG
        vc.font          = F_BOLD if total_row else F_NORMAL
        vc.alignment     = _right()
        vc.border        = _border("thin")

        # Note / Ghi chú (cột G-J)
        if note:
            ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=10)
            nc = ws.cell(row=row, column=7, value=note)
            nc.font      = F_SMALL
            nc.alignment = Alignment(vertical="center")
            nc.border    = _border("thin")

        # Màu nền nếu có / Background color
        bg = highlight or (C_TOTAL_BG if total_row else None)
        if bg:
            for c in range(1, 11):
                ws.cell(row=row, column=c).fill = _fill(bg)

        row += 1

    def write_section_header(title):
        nonlocal row
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        h = ws.cell(row=row, column=1, value=f"  {title}")
        h.font      = Font(name="Calibri", bold=True, size=10, color="1F4E79")
        h.fill      = _fill(C_SECTION_BG)
        h.alignment = Alignment(vertical="center")
        h.border    = _border("medium")
        row += 1

    # ── Section 1: Opening Balance ────────────────────────────────────────────
    write_section_header("SECTION 1 – OPENING BALANCE / Số dư đầu kỳ")
    write_row("Opening Stock (Certified) / Tồn kho đầu kỳ",
              float(result.opening_stock),
              "kg tồn từ kỳ trước / kg carried over from previous period")

    # ── Section 2: Inputs ────────────────────────────────────────────────────
    write_section_header("SECTION 2 – CERTIFIED MATERIAL PURCHASED / Mua vật liệu có chứng nhận")
    write_row("Certified Material Purchased (∑ In TCs) / Mua vật liệu có chứng nhận",
              float(result.certified_purchased),
              f"{result.number_tc_in} In TCs received in period")
    write_row("Total Available / Tổng khả dụng",
              float(result.total_available),
              "= Opening + Purchased",
              total_row=True)

    # ── Section 3: Production ────────────────────────────────────────────────
    write_section_header("SECTION 3 – PRODUCTION LOSS / Hao hụt sản xuất")
    write_row("Issued to Production / Xuất sản xuất",
              float(result.issued_to_production),
              "= Tổng xuất cắt / Total issued to cutting")

    if result.cutting:
        write_row(f"  Cutting Loss / Hao hụt Cắt  ({result.cutting.loss_pct:.2f}%)",
                  float(result.cutting.loss_qty),
                  f"Input: {result.cutting.input_qty:.3f} kg → Output: {result.cutting.output_qty:.3f} kg")

    if result.sewing:
        write_row(f"  Sewing Loss / Hao hụt May  ({result.sewing.loss_pct:.2f}%)",
                  float(result.sewing.loss_qty),
                  f"Input: {result.sewing.input_qty:.3f} kg → Output: {result.sewing.output_qty:.3f} kg")

    if result.packing:
        write_row(f"  Packing Loss / Hao hụt Đóng gói  ({result.packing.loss_pct:.2f}%)",
                  float(result.packing.loss_qty),
                  f"Input: {result.packing.input_qty:.3f} kg → Output: {result.packing.output_qty:.3f} kg")

    write_row(f"Total Process Loss / Tổng hao hụt  ({result.total_loss_pct:.2f}%)",
              float(result.total_process_loss),
              f"Overall yield: {result.overall_yield_pct:.2f}%",
              total_row=True)
    write_row("Certified Produced / Thành phẩm có chứng nhận",
              float(result.certified_produced), "= Đầu ra công đoạn cuối")

    # ── Section 4: Outputs ───────────────────────────────────────────────────
    write_section_header("SECTION 4 – CERTIFIED PRODUCT SOLD / Sản phẩm có chứng nhận đã bán")
    write_row("Certified Sold (∑ Out TCs) / Đã bán có chứng nhận",
              float(result.certified_sold),
              f"{result.number_tc_out} Out TCs issued in period")

    # ── Section 5: Closing Balance ────────────────────────────────────────────
    write_section_header("SECTION 5 – CLOSING BALANCE / Số dư cuối kỳ")
    write_row("Closing Stock – Raw Material / Tồn nguyên liệu",
              float(result.closing_stock_raw))
    write_row("Closing Stock – Finished Goods / Tồn thành phẩm",
              float(result.closing_stock_fg))
    write_row("Closing Stock TOTAL / Tổng tồn kho",
              float(result.closing_stock_total),
              "= Raw + Finished Goods",
              total_row=True)

    # ── Section 6: Reconciliation ────────────────────────────────────────────
    write_section_header("SECTION 6 – MASS BALANCE RECONCILIATION / Đối soát Mass Balance")
    write_row("Theoretical Closing / Số dư lý thuyết",
              float(result.theoretical_closing),
              "= Opening + Purchased − Sold − Total Loss")
    write_row("Actual Closing / Số dư thực tế",
              float(result.closing_stock_total))

    # Dòng chênh lệch – tô màu theo status
    diff_color = {"PASS": C_PASS, "WARNING": C_WARNING, "FAIL": C_FAIL}.get(result.balance_status, "FFFFFF")
    write_row(
        f"Difference / Chênh lệch  ({'+' if result.difference >= 0 else ''}{result.difference_pct:.4f}%)",
        float(result.difference),
        "= Actual − Theoretical  |  + surplus, − deficit",
        highlight=diff_color,
    )

    # Dòng STATUS to lớn
    nonlocal_row = row
    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=10)
    sc = ws.cell(row=row, column=1,
                 value=f"BALANCE STATUS:  {result.balance_status}  "
                       f"({'✓ COMPLIANT – chênh lệch trong ngưỡng cho phép' if result.balance_status == 'PASS' else '⚠ WARNING – cần xem xét trước nộp kiểm toán' if result.balance_status == 'WARNING' else '✗ NON-COMPLIANT – vượt ngưỡng, yêu cầu điều tra'})")
    sc.font      = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    sc.fill      = _fill(diff_color)
    sc.alignment = _center()
    sc.border    = _border("medium")
    row += 2

    # Alerts
    if result.alerts or result.warnings:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.cell(row=row, column=1, value="ALERTS & WARNINGS / CẢNH BÁO:").font = F_BOLD
        row += 1
        for alert in result.alerts + result.warnings:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            ws.cell(row=row, column=1, value=f"  ⚠  {alert}").font = F_SMALL
            row += 1

    # ── TC Details ───────────────────────────────────────────────────────────
    if result.tc_in_details:
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="IN TCs RECEIVED IN PERIOD / IN TC nhận trong kỳ:").font = F_BOLD
        row += 1
        _write_header_row(ws, row,
                          ["TC Number", "Issue Date", "Supplier", "Quantity (kg)", "Standard", "Notes"],
                          [25, 14, 30, 16, 10, 25])
        row += 1
        for i, tc in enumerate(result.tc_in_details):
            bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
            for col, val in enumerate([
                tc["tc_number"], tc["issue_date"], tc["supplier"],
                tc["quantity_kg"], tc["standard"], ""
            ], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill   = _fill(bg)
                cell.border = _border("thin")
                if col == 4:
                    cell.number_format = "#,##0.000"
                    cell.alignment = _right()
            row += 1

    if result.tc_out_details:
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="OUT TCs ISSUED IN PERIOD / OUT TC phát hành trong kỳ:").font = F_BOLD
        row += 1
        _write_header_row(ws, row,
                          ["TC Number", "Issue Date", "Buyer", "Quantity (kg)", "Product", "Standard"],
                          [25, 14, 30, 16, 30, 10])
        row += 1
        for i, tc in enumerate(result.tc_out_details):
            bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
            for col, val in enumerate([
                tc["tc_number"], tc["issue_date"], tc["buyer"],
                tc["quantity_kg"], tc["product"], ""
            ], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill   = _fill(bg)
                cell.border = _border("thin")
                if col == 4:
                    cell.number_format = "#,##0.000"
                    cell.alignment = _right()
            row += 1

    # Cột rộng / Column widths
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["D"].width = 20


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Monthly Breakdown Table
# ─────────────────────────────────────────────────────────────────────────────

def write_monthly_breakdown(ws, rows: List[dict]) -> None:
    """Bảng breakdown theo tháng / Monthly breakdown table"""
    ws.title = "Monthly_MB"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:P1")
    ws.cell(row=1, column=1, value="MONTHLY MASS BALANCE BREAKDOWN / PHÂN TÍCH MASS BALANCE THEO THÁNG").font = \
        Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws.cell(row=1, column=1).alignment = _center()

    headers = [
        "Month", "Opening (kg)",
        "Purchased (kg)", "# In TCs",
        "Issued to Prod (kg)",
        "Cut Loss (kg)", "Cut Loss %",
        "Sew Loss (kg)", "Sew Loss %",
        "Pack Loss (kg)", "Pack Loss %",
        "Total Loss (kg)", "Total Loss %",
        "Sold (kg)", "# Out TCs",
        "Closing (kg)", "Status",
    ]
    widths = [12, 14, 14, 9, 16, 13, 11, 12, 11, 13, 11, 13, 12, 13, 9, 13, 10]
    _write_header_row(ws, 3, headers, widths)

    totals = {k: Decimal("0") for k in ["opening", "purchased", "issued", "cut_loss",
                                         "sew_loss", "pack_loss", "total_loss", "sold", "closing"]}
    for i, row_data in enumerate(rows, 4):
        if "error" in row_data:
            ws.cell(row=i, column=1, value=row_data.get("month_name", "")).font = F_BOLD
            ws.cell(row=i, column=2, value=f"ERROR: {row_data['error']}").font = \
                Font(name="Calibri", color="FF0000")
            continue

        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        status = row_data.get("balance_status", "PENDING")
        status_bg = {"PASS": C_PASS, "WARNING": C_WARNING, "FAIL": C_FAIL}.get(status, bg)

        vals = [
            row_data.get("month_name", ""),
            row_data.get("opening_stock_kg", 0),
            row_data.get("certified_purchased_kg", 0),
            row_data.get("number_tc_in", 0),
            row_data.get("issued_to_production_kg", 0),
            (row_data.get("cutting") or {}).get("loss_kg", 0),
            (row_data.get("cutting") or {}).get("loss_pct", 0),
            (row_data.get("sewing")  or {}).get("loss_kg", 0),
            (row_data.get("sewing")  or {}).get("loss_pct", 0),
            (row_data.get("packing") or {}).get("loss_kg", 0),
            (row_data.get("packing") or {}).get("loss_pct", 0),
            row_data.get("total_process_loss_kg", 0),
            row_data.get("total_loss_pct", 0),
            row_data.get("certified_sold_kg", 0),
            row_data.get("number_tc_out", 0),
            row_data.get("closing_stock_total_kg", 0),
            status,
        ]
        pct_cols = {7, 9, 11, 13}   # Cột % / PCT columns (1-indexed)
        kg_cols  = {2, 3, 5, 6, 8, 10, 12, 14, 16}

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill   = _fill(status_bg if col == 17 else bg)
            cell.border = _border("thin")
            if col in pct_cols:
                cell.number_format = "0.00%"
                cell.value = (val / 100) if val else 0   # openpyxl % format
                cell.alignment = _right()
            elif col in kg_cols:
                cell.number_format = "#,##0.000"
                cell.alignment = _right()
            elif col == 1:
                cell.font = F_BOLD

    # Totals row
    total_row = len(rows) + 4
    ws.cell(row=total_row, column=1, value="TOTAL / TỔNG CỘNG").font = F_BOLD
    for cell in ws[total_row]:
        cell.fill   = _fill(C_TOTAL_BG)
        cell.border = _border("medium")


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: Summary Maintain Stock
# ─────────────────────────────────────────────────────────────────────────────

def write_summary_stock(ws, rows: List[StockSummaryRow], period_start: date, period_end: date) -> None:
    """Summary tồn kho cuối kỳ / End-of-period stock summary"""
    ws.title = "Summary_Stock"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1,
            value=f"SUMMARY MAINTAIN STOCK / TÓM TẮT TỒN KHO  –  {period_start} → {period_end}").font = \
        Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws.cell(row=1, column=1).alignment = _center()

    headers = ["Material Code", "Material Name", "Opening (kg)",
               "Received (kg)", "Issued to Prod (kg)",
               "Sold TC Out (kg)", "Closing (kg)", "In TC Count", "Status"]
    widths  = [16, 30, 14, 14, 18, 16, 14, 12, 14]
    _write_header_row(ws, 3, headers, widths)

    for i, r in enumerate(rows, 4):
        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        status_bg = C_PASS if r.status == "OK" else C_FAIL
        vals = [r.material_code, r.material_name,
                float(r.opening_stock), float(r.received_kg),
                float(r.issued_production), float(r.sold_kg),
                float(r.closing_stock), r.tc_in_count, r.status]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill   = _fill(status_bg if col == 9 else bg)
            cell.border = _border("thin")
            if col in {3, 4, 5, 6, 7}:
                cell.number_format = "#,##0.000"
                cell.alignment = _right()

    # Grand total
    total_row = len(rows) + 4
    ws.cell(row=total_row, column=1, value="TOTAL").font = F_BOLD
    for col in {3, 4, 5, 6, 7}:
        vals = [ws.cell(row=i, column=col).value for i in range(4, total_row) if isinstance(ws.cell(row=i, column=col).value, (int, float))]
        total_cell = ws.cell(row=total_row, column=col, value=sum(vals))
        total_cell.number_format = "#,##0.000"
        total_cell.alignment = _right()
        total_cell.font = F_BOLD
        total_cell.fill = _fill(C_TOTAL_BG)
        total_cell.border = _border("medium")


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: Stock Card
# ─────────────────────────────────────────────────────────────────────────────

def write_stock_card(ws, sc: StockCardResult) -> None:
    """Thẻ kho chi tiết / Detailed stock card"""
    ws.title = "Stock_Card"
    ws.sheet_view.showGridLines = False

    row = 1
    ws.merge_cells(f"A{row}:J{row}")
    ws.cell(row=row, column=1,
            value=f"STOCK CARD / THẺ KHO  –  {sc.material_code} – {sc.material_name}").font = \
        Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws.cell(row=row, column=1).alignment = _center()

    row = 2
    ws.merge_cells(f"A{row}:J{row}")
    ws.cell(row=row, column=1,
            value=f"Period: {sc.date_from} → {sc.date_to}  |  "
                  f"Opening: {sc.opening_balance:.3f} kg  |  "
                  f"Closing: {sc.closing_balance:.3f} kg").alignment = _center()

    headers = ["Date", "Transaction Type", "Label", "In TC #", "Out TC #",
               "Batch #", "Reference", "In (kg)", "Out (kg)", "Balance (kg)"]
    widths  = [12, 18, 30, 18, 18, 14, 14, 12, 12, 14]
    _write_header_row(ws, 4, headers, widths)

    row = 5
    for i, line in enumerate(sc.lines):
        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        vals = [
            line.date, line.transaction_type, line.label,
            line.tc_in_number or "", line.tc_out_number or "",
            line.batch_number or "", line.reference_number or "",
            float(line.qty_in)  if line.qty_in  is not None else "",
            float(line.qty_out) if line.qty_out is not None else "",
            float(line.running_balance),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = _fill(bg)
            cell.border = _border("thin")
            if col in {8, 9, 10}:
                if val != "":
                    cell.number_format = "#,##0.000"
                    cell.alignment     = _right()
                    if col == 10 and isinstance(val, float) and val < 0:
                        cell.font = Font(name="Calibri", size=10, color="FF0000")
        row += 1

    # Totals
    ws.cell(row=row, column=7, value="TOTAL").font = F_BOLD
    ws.cell(row=row, column=8, value=float(sc.total_in)).number_format  = "#,##0.000"
    ws.cell(row=row, column=9, value=float(sc.total_out)).number_format = "#,##0.000"
    ws.cell(row=row, column=10, value=float(sc.closing_balance)).number_format = "#,##0.000"
    for col in {7, 8, 9, 10}:
        ws.cell(row=row, column=col).fill   = _fill(C_TOTAL_BG)
        ws.cell(row=row, column=col).border = _border("medium")
        ws.cell(row=row, column=col).font   = F_BOLD
        ws.cell(row=row, column=col).alignment = _right()


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: Traceability
# ─────────────────────────────────────────────────────────────────────────────

def write_traceability(ws, result: TraceabilityResult) -> None:
    """Báo cáo truy xuất nguồn gốc / Traceability report"""
    ws.title = "Traceability"
    ws.sheet_view.showGridLines = False

    row = 1
    ws.merge_cells(f"A{row}:J{row}")
    ws.cell(row=row, column=1,
            value=f"TRACEABILITY REPORT / BÁO CÁO TRUY XUẤT  –  {result.tc_in_number}").font = \
        Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws.cell(row=row, column=1).alignment = _center()

    row = 3
    info_rows = [
        ("In TC Number",           result.tc_in_number),
        ("Issue Date",             result.tc_in_issue_date),
        ("Supplier",               result.supplier_name),
        ("Material",               f"{result.material_code} – {result.material_name}"),
        ("Certification Standard", result.certification_standard),
        ("Certified Quantity",     f"{result.certified_qty:.3f} kg"),
    ]
    for label, val in info_rows:
        ws.cell(row=row, column=1, value=label).font = F_BOLD
        ws.cell(row=row, column=3, value=val)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1, value="PRODUCTION BATCHES / LÔ SẢN XUẤT").font = \
        Font(name="Calibri", bold=True, size=11, color="1F4E79")
    row += 1

    headers = ["Batch #", "Product", "Order #", "Period", "Status",
               "Input (kg)", "Output (kg)", "Loss (kg)", "Loss %",
               "Cut %", "Sew %", "Pack %"]
    widths  = [16, 25, 14, 22, 12, 12, 12, 11, 10, 9, 9, 9]
    _write_header_row(ws, row, headers, widths)
    row += 1

    for i, b in enumerate(result.batches):
        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        cut_s = next((s for s in b.stages if s.stage == "cutting"), None)
        sew_s = next((s for s in b.stages if s.stage == "sewing"),  None)
        pack_s= next((s for s in b.stages if s.stage == "packing"), None)

        vals = [
            b.batch_number, b.product_name, b.order_number or "",
            f"{b.start_date} → {b.end_date or 'ongoing'}", b.status,
            float(b.actual_in)  if b.actual_in  else "",
            float(b.actual_out) if b.actual_out else "",
            float(b.total_loss_kg)  if b.total_loss_kg  else "",
            float(b.total_loss_pct) / 100 if b.total_loss_pct else "",
            float(cut_s.loss_pct) / 100 if cut_s  else "",
            float(sew_s.loss_pct) / 100 if sew_s  else "",
            float(pack_s.loss_pct)/ 100 if pack_s else "",
        ]
        pct_cols = {9, 10, 11, 12}
        kg_cols  = {6, 7, 8}
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = _fill(bg)
            cell.border = _border("thin")
            if col in pct_cols and val != "":
                cell.number_format = "0.00%"
                cell.alignment = _right()
            elif col in kg_cols and val != "":
                cell.number_format = "#,##0.000"
                cell.alignment = _right()
        row += 1

    # Out TCs
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="LINKED OUT TCs / OUT TC LIÊN KẾT").font = \
        Font(name="Calibri", bold=True, size=11, color="1F4E79")
    row += 1

    headers2 = ["Out TC #", "Issue Date", "Buyer", "Product", "Allocated (kg)", "Alloc %", "Status"]
    widths2  = [22, 14, 28, 28, 15, 12, 12]
    _write_header_row(ws, row, headers2, widths2)
    row += 1

    for i, link in enumerate(result.linked_tc_outs):
        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        vals = [link.tc_out_number, link.issue_date, link.buyer_name, link.product_name,
                float(link.allocated_qty),
                float(link.allocation_pct) / 100 if link.allocation_pct else "",
                link.status]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = _fill(bg)
            cell.border = _border("thin")
            if col == 5:
                cell.number_format = "#,##0.000"
                cell.alignment = _right()
            elif col == 6 and val != "":
                cell.number_format = "0.00%"
                cell.alignment = _right()
        row += 1

    # Summary box
    row += 2
    summary_data = [
        ("Certified Quantity",    float(result.certified_qty)),
        ("Issued to Production",  float(result.total_issued_to_prod)),
        ("Total Produced",        float(result.total_produced)),
        ("Process Loss",          f"{result.total_process_loss_kg:.3f} kg ({result.total_process_loss_pct:.2f}%)"),
        ("Sold with Out TC",      float(result.total_sold_with_tc)),
        ("Remaining in Stock",    float(result.remaining_in_stock)),
        ("Utilization",           f"{result.utilization_pct:.2f}%"),
        ("Balance OK",            "YES ✓" if result.is_balanced else "NO ✗ – OVERCLAIM"),
    ]
    for label, val in summary_data:
        ws.cell(row=row, column=1, value=label).font = F_BOLD
        c = ws.cell(row=row, column=3, value=val)
        if isinstance(val, float):
            c.number_format = "#,##0.000"
            c.alignment = _right()
        row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 6: Loss Analysis
# ─────────────────────────────────────────────────────────────────────────────

def write_loss_analysis(ws, result: LossAnalysisResult) -> None:
    """Báo cáo phân tích hao hụt / Loss analysis report"""
    ws.title = "Loss_Analysis"
    ws.sheet_view.showGridLines = False

    row = 1
    ws.merge_cells(f"A{row}:J{row}")
    ws.cell(row=row, column=1,
            value=f"LOSS ANALYSIS REPORT / PHÂN TÍCH HAO HỤT  –  {result.period_start} → {result.period_end}").font = \
        Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws.cell(row=row, column=1).alignment = _center()

    # Stage summary
    row = 3
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="STAGE SUMMARY / TÓM TẮT THEO CÔNG ĐOẠN").font = F_BOLD
    row += 1
    headers = ["Stage", "Total Input (kg)", "Total Output (kg)", "Total Loss (kg)",
               "Avg Loss %", "Min Loss %", "Max Loss %", "# Batches", "Above Threshold?"]
    widths  = [14, 16, 16, 15, 12, 12, 12, 11, 18]
    _write_header_row(ws, row, headers, widths)
    row += 1

    for s in result.stage_summaries:
        bg = C_FAIL if s.above_threshold else C_PASS
        vals = [s.stage.upper(), float(s.total_input_kg), float(s.total_output_kg),
                float(s.total_loss_kg),
                float(s.avg_loss_pct) / 100, float(s.min_loss_pct) / 100,
                float(s.max_loss_pct) / 100, s.num_batches,
                f"YES (>{s.threshold_pct}%)" if s.above_threshold else "OK"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = _fill(bg)
            cell.border = _border("thin")
            if col in {2, 3, 4}:
                cell.number_format = "#,##0.000"
                cell.alignment = _right()
            elif col in {5, 6, 7}:
                cell.number_format = "0.00%"
                cell.alignment = _right()
        row += 1

    # Top loss batches
    row += 2
    ws.merge_cells(f"A{row}:I{row}")
    ws.cell(row=row, column=1, value="TOP 10 HIGHEST LOSS BATCHES / 10 LÔ HAO HỤT CAO NHẤT").font = F_BOLD
    row += 1
    headers2 = ["Batch #", "Product", "Order #", "In TC #",
                "Input (kg)", "Output (kg)", "Loss (kg)", "Loss %", "Status"]
    widths2  = [16, 25, 14, 18, 13, 13, 12, 11, 12]
    _write_header_row(ws, row, headers2, widths2)
    row += 1

    for i, b in enumerate(result.top_loss_batches):
        bg = C_FAIL if b.total_loss_pct > Decimal("20") else (
            C_WARNING if b.total_loss_pct > Decimal("10") else C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        )
        vals = [b.batch_number, b.product_name, b.order_number or "",
                b.tc_in_number or "",
                float(b.total_input_kg), float(b.total_output_kg), float(b.total_loss_kg),
                float(b.total_loss_pct) / 100, b.status]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = _fill(bg)
            cell.border = _border("thin")
            if col in {5, 6, 7}:
                cell.number_format = "#,##0.000"
                cell.alignment = _right()
            elif col == 8:
                cell.number_format = "0.00%"
                cell.alignment = _right()
        row += 1

    # Monthly trend
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1, value="MONTHLY TREND / XU HƯỚNG THÁNG").font = F_BOLD
    row += 1
    headers3 = ["Month", "Total Input (kg)", "Total Loss (kg)", "Avg Loss %",
                "Cutting Loss %", "Sewing Loss %", "Packing Loss %", "# Batches"]
    _write_header_row(ws, row, headers3, [12, 16, 15, 13, 14, 13, 14, 11])
    row += 1

    for m in result.monthly_trend:
        bg = C_FAIL if m.avg_loss_pct > Decimal("20") else (
            C_WARNING if m.avg_loss_pct > Decimal("10") else "FFFFFF"
        )
        vals = [m.year_month,
                float(m.total_input_kg), float(m.total_loss_kg),
                float(m.avg_loss_pct) / 100, float(m.cutting_loss_pct) / 100,
                float(m.sewing_loss_pct) / 100, float(m.packing_loss_pct) / 100,
                m.num_batches]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = _fill(bg)
            cell.border = _border("thin")
            if col in {2, 3}:
                cell.number_format = "#,##0.000"
                cell.alignment = _right()
            elif col in {4, 5, 6, 7}:
                cell.number_format = "0.00%"
                cell.alignment = _right()
        row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Main ExcelExporter class
# ─────────────────────────────────────────────────────────────────────────────

class ExcelExporter:
    """
    Xuất tất cả báo cáo ra một file Excel duy nhất
    Export all reports to a single Excel workbook

    Usage:
        exporter = ExcelExporter()
        buffer = exporter.export_full_workbook(
            annual_mb     = annual_result,
            monthly_rows  = monthly_rows,
            summary_rows  = summary_rows,
            stock_card    = stock_card_result,
            traceability  = trace_result,
            loss_analysis = loss_result,
            company_name  = "ABC Garment Factory",
        )
        with open("GRS_Mass_Balance_2024.xlsx", "wb") as f:
            f.write(buffer.getvalue())
    """

    def export_full_workbook(
        self,
        annual_mb      : Optional[MassBalanceResult]   = None,
        monthly_rows   : Optional[List[dict]]          = None,
        summary_rows   : Optional[List[StockSummaryRow]] = None,
        stock_card     : Optional[StockCardResult]     = None,
        traceability   : Optional[TraceabilityResult]  = None,
        loss_analysis  : Optional[LossAnalysisResult]  = None,
        company_name   : str                           = "",
        period_start   : Optional[date]                = None,
        period_end     : Optional[date]                = None,
    ) -> io.BytesIO:
        """
        Tạo workbook Excel với tất cả sheets
        Create Excel workbook with all report sheets
        """
        wb = Workbook()
        wb.remove(wb.active)   # Xóa sheet mặc định

        # ── Thêm từng sheet ────────────────────────────────────────────────
        if annual_mb:
            ws = wb.create_sheet()
            write_annual_mass_balance(ws, annual_mb, company_name)

        if monthly_rows:
            ws = wb.create_sheet()
            write_monthly_breakdown(ws, monthly_rows)

        if summary_rows and period_start and period_end:
            ws = wb.create_sheet()
            write_summary_stock(ws, summary_rows, period_start, period_end)

        if stock_card:
            ws = wb.create_sheet()
            write_stock_card(ws, stock_card)

        if traceability:
            ws = wb.create_sheet()
            write_traceability(ws, traceability)

        if loss_analysis:
            ws = wb.create_sheet()
            write_loss_analysis(ws, loss_analysis)

        # ── Sheet thông tin / Info sheet ───────────────────────────────────
        ws_info = wb.create_sheet("Report_Info")
        ws_info["A1"] = "GRS/GOTS Mass Balance & Traceability System"
        ws_info["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        ws_info["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_info["A3"] = f"Company: {company_name}"
        ws_info["A4"] = "Powered by Volume Reconciliation & Material Traceability System"
        ws_info.column_dimensions["A"].width = 60

        # ── Freeze pane và print settings cho mỗi sheet ───────────────────
        for ws in wb.worksheets:
            ws.freeze_panes = "A5"
            ws.print_title_rows = "1:4"
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage   = True

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_annual_only(
        self,
        result       : MassBalanceResult,
        company_name : str = "",
    ) -> io.BytesIO:
        """Xuất chỉ Annual Mass Balance / Export only Annual Mass Balance"""
        return self.export_full_workbook(annual_mb=result, company_name=company_name)
